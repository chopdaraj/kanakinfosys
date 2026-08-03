from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import io
import logging
import secrets
import string
from datetime import datetime, timezone, timedelta
from typing import Optional, List

import asyncio
import bcrypt
import jwt
import pandas as pd
from bson import ObjectId
from fastapi import FastAPI, APIRouter, Depends, HTTPException, Request, Response
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, EmailStr, Field
from starlette.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware


# -------------------- Configuration --------------------
JWT_SECRET = os.environ["JWT_SECRET"]
JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_MIN = 60 * 24 * 7  # 7 days for MVP simplicity

DAILY_RATE = 0.03 / 30.0        # 3% per month distributed daily
LOCK_MONTHS = 6
MIN_DEPOSIT = 100000            # 1 lakh

mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

app = FastAPI(title="Kanak Infosys API")
api = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("kanak")


# -------------------- Helpers --------------------
def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()


def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False


def create_access_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_MIN),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def gen_referral_code(n: int = 8) -> str:
    alpha = string.ascii_uppercase + string.digits
    return "KNK" + "".join(secrets.choice(alpha) for _ in range(n))


def serialize_user(u: dict) -> dict:
    return {
        "id": str(u["_id"]),
        "name": u.get("name"),
        "email": u.get("email"),
        "role": u.get("role", "client"),
        "referral_code": u.get("referral_code"),
        "referred_by_code": u.get("referred_by_code"),
        "serial_number": u.get("serial_number"),
        "client_id": u.get("client_id"),
        "address_line1": u.get("address_line1", ""),
        "address_line2": u.get("address_line2", ""),
        "bank_name": u.get("bank_name", ""),
        "account_holder": u.get("account_holder", ""),
        "account_number": u.get("account_number", ""),
        "ifsc": u.get("ifsc", ""),
        "nominee": u.get("nominee", ""),
        "profile_photo": u.get("profile_photo", ""),
        "aadhaar_number": u.get("aadhaar_number", ""),
        "pan_number": u.get("pan_number", ""),
        "aadhaar_front": u.get("aadhaar_front", ""),
        "aadhaar_back": u.get("aadhaar_back", ""),
        "pan_front": u.get("pan_front", ""),
        "pan_back": u.get("pan_back", ""),
        "kyc_status": u.get("kyc_status", "not_started"),
        "dob": u.get("dob", ""),
        "city": u.get("city", ""),
        "state": u.get("state", ""),
        "pincode": u.get("pincode", ""),
        "branch_name": u.get("branch_name", ""),
        "selfie": u.get("selfie", ""),
        "cancelled_cheque": u.get("cancelled_cheque", ""),
        "phone": u.get("phone", ""),
        "email_verified": u.get("email_verified", False),
        "mobile_verified": u.get("mobile_verified", False),
        "created_at": u.get("created_at").isoformat() if u.get("created_at") else None,
    }



async def get_current_user(request: Request) -> dict:
    token = None
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        token = auth[7:]
    if not token:
        token = request.cookies.get("access_token")
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user


async def require_admin(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user


async def user_total_deposit(user_id: ObjectId) -> float:
    cursor = db.deposits.find({"user_id": user_id, "status": "approved"})
    total = 0.0
    async for d in cursor:
        total += float(d.get("amount", 0))
    return total


async def compute_user_balances(user_id: ObjectId) -> dict:
    principal = 0.0
    locked_amount = 0.0
    unlocked_amount = 0.0
    pending_deposits = 0.0
    now = datetime.now(timezone.utc)
    
    cfg = await db.settings.find_one({"_id": "global_config"})
    lock_months = int(cfg.get("lock_period_months", LOCK_MONTHS)) if cfg else LOCK_MONTHS
    
    async for d in db.deposits.find({"user_id": user_id}):
        amt = float(d.get("amount", 0))
        if d.get("status") == "approved":
            principal += amt
            lu = d.get("lock_until")
            if isinstance(lu, str):
                lu = datetime.fromisoformat(lu)
            if lu and lu.tzinfo is None:
                lu = lu.replace(tzinfo=timezone.utc)
            if lu and lu <= now:
                unlocked_amount += amt
            else:
                locked_amount += amt
        elif d.get("status") == "pending":
            pending_deposits += amt
            
    total_profit_earned = 0.0
    today_profit_earning = 0.0
    today = datetime.now(timezone.utc).date()
    async for p in db.profit_credits.find({"user_id": user_id}):
        amt = float(p["amount"])
        total_profit_earned += amt
        dt = p["credited_at"]
        if isinstance(dt, str):
            dt = datetime.fromisoformat(dt)
        if dt.date() == today:
            today_profit_earning += amt
            
    total_referral_earned = 0.0
    today_referral_earning = 0.0
    async for r in db.referral_credits.find({"user_id": user_id}):
        amt = float(r["amount"])
        total_referral_earned += amt
        dt = r["credited_at"]
        if isinstance(dt, str):
            dt = datetime.fromisoformat(dt)
        if dt.date() == today:
            today_referral_earning += amt
            
    total_withdrawn = 0.0
    pending_withdrawals = 0.0
    async for w in db.withdrawals.find({"user_id": user_id}):
        amt = float(w["amount"])
        if w.get("status") in ["approved", "paid"]:
            total_withdrawn += amt
        elif w.get("status") in ["pending", "under_review"]:
            pending_withdrawals += amt

            
    # Monthly credits (profit + referral commissions in the current calendar month)
    now = datetime.now(timezone.utc)
    month_start = datetime(now.year, now.month, 1, tzinfo=timezone.utc)
    
    monthly_profit_earned = 0.0
    async for p in db.profit_credits.find({"user_id": user_id, "credited_at": {"$gte": month_start}}):
        monthly_profit_earned += float(p["amount"])
        
    monthly_referral_earned = 0.0
    async for r in db.referral_credits.find({"user_id": user_id, "credited_at": {"$gte": month_start}}):
        monthly_referral_earned += float(r["amount"])
        
    monthly_credits = monthly_profit_earned + monthly_referral_earned
    
    user = await db.users.find_one({"_id": user_id})
    referral_code = user.get("referral_code", "") if user else ""
    referral_count = await db.users.count_documents({"referred_by_code": referral_code}) if referral_code else 0
    
    total_earned = total_profit_earned + total_referral_earned
    today_earning = today_profit_earning + today_referral_earning
    withdrawable_amount = max(0.0, unlocked_amount + total_earned - total_withdrawn - pending_withdrawals)
    roi_percent = round((total_earned / principal * 100), 2) if principal > 0 else 0.0
    
    return {
        "principal": round(principal, 2),
        "locked_amount": round(locked_amount, 2),
        "unlocked_amount": round(unlocked_amount, 2),
        "pending_deposits": round(pending_deposits, 2),
        "total_profit_earned": round(total_profit_earned, 2),
        "total_referral_earned": round(total_referral_earned, 2),
        "total_earned": round(total_earned, 2),
        "today_earning": round(today_earning, 2),
        "monthly_credits": round(monthly_credits, 2),
        "total_withdrawn": round(total_withdrawn, 2),
        "pending_withdrawals": round(pending_withdrawals, 2),
        "referral_count": referral_count,
        "withdrawable_amount": round(withdrawable_amount, 2),
        "net_profit": round(total_earned, 2),
        "roi_percent": roi_percent,
        "lock_period": f"{lock_months} Months",
    }


# -------------------- Models --------------------
class RegisterIn(BaseModel):
    name: str
    email: EmailStr
    password: str = Field(min_length=6)
    referral_code: Optional[str] = None


class LoginIn(BaseModel):
    email: EmailStr
    password: str


class ChangePasswordIn(BaseModel):
    new_password: str = Field(min_length=6)


class ProfileIn(BaseModel):
    address_line1: Optional[str] = ""
    address_line2: Optional[str] = ""
    email: Optional[EmailStr] = None
    phone: Optional[str] = ""
    bank_name: Optional[str] = ""
    account_holder: Optional[str] = ""
    account_number: Optional[str] = ""
    ifsc: Optional[str] = ""
    nominee: Optional[str] = ""
    profile_photo: Optional[str] = ""
    aadhaar_number: Optional[str] = ""
    pan_number: Optional[str] = ""
    aadhaar_front: Optional[str] = ""  # base64 data URL
    aadhaar_back: Optional[str] = ""
    pan_front: Optional[str] = ""
    pan_back: Optional[str] = ""
    dob: Optional[str] = ""
    city: Optional[str] = ""
    state: Optional[str] = ""
    pincode: Optional[str] = ""
    branch_name: Optional[str] = ""
    selfie: Optional[str] = ""
    cancelled_cheque: Optional[str] = ""
    phone: Optional[str] = ""


class AdminEditClientIn(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    address_line1: Optional[str] = None
    address_line2: Optional[str] = None
    bank_name: Optional[str] = None
    account_holder: Optional[str] = None
    account_number: Optional[str] = None
    ifsc: Optional[str] = None
    nominee: Optional[str] = None
    aadhaar_number: Optional[str] = None
    pan_number: Optional[str] = None
    dob: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    branch_name: Optional[str] = None
    phone: Optional[str] = None



class DepositIn(BaseModel):
    amount: float = Field(gt=0)
    payment_screenshot: str
    payment_method: str
    transaction_id: str
    deposit_date: str
    remarks: Optional[str] = ""


class BroadcastIn(BaseModel):
    title: str
    description: str


class CreditProfitIn(BaseModel):
    amount: float = Field(gt=0)
    note: Optional[str] = ""


class BulkCreditIn(BaseModel):
    rate_percent: Optional[float] = None  # e.g. 0.1 for 0.10% of principal
    flat_amount: Optional[float] = None
    note: Optional[str] = ""
    client_ids: Optional[List[str]] = None


class WithdrawIn(BaseModel):
    amount: float = Field(gt=0)


class KycStatusIn(BaseModel):
    status: str  # "verified" | "rejected" | "pending"


class SettingsIn(BaseModel):
    min_deposit: Optional[float] = None
    profit_percentage_daily: Optional[float] = None
    referral_percentage_monthly: Optional[float] = None
    lock_period_months: Optional[int] = None
    company_bank_name: Optional[str] = None
    company_account_number: Optional[str] = None
    company_account_holder: Optional[str] = None
    company_ifsc: Optional[str] = None
    company_branch: Optional[str] = None
    company_email: Optional[str] = None
    company_whatsapp: Optional[str] = None
    company_sms: Optional[str] = None
    maintenance_mode: Optional[bool] = None




class SendEmailOtpIn(BaseModel):
    email: EmailStr

class VerifyEmailOtpIn(BaseModel):
    email: EmailStr
    otp: str

class SendMobileOtpIn(BaseModel):
    phone: str

class VerifyMobileOtpIn(BaseModel):
    phone: str
    otp: str


# -------------------- Auth Routes --------------------
@api.post("/auth/register")
async def register(payload: RegisterIn, response: Response):
    email = payload.email.lower()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")

    referred_by = None
    if payload.referral_code:
        ref_code = payload.referral_code.strip().upper()
        parent = await db.users.find_one({"referral_code": ref_code})
        if not parent:
            raise HTTPException(status_code=400, detail="Invalid referral code")
        referred_by = ref_code

    # Allocate unique permanent sequence numbers
    max_user = await db.users.find_one({"role": "client"}, sort=[("serial_number", -1)])
    next_serial = (max_user["serial_number"] + 1) if max_user and "serial_number" in max_user else 50

    while True:
        client_id = f"CLI{next_serial:06d}"
        code = f"KNK{next_serial:04d}"
        if not await db.users.find_one({"$or": [{"serial_number": next_serial}, {"client_id": client_id}, {"referral_code": code}]}):
            break
        next_serial += 1

    otp_code = "".join(secrets.choice(string.digits) for _ in range(6))
    otp_expiry = datetime.now(timezone.utc) + timedelta(minutes=10)

    doc = {
        "name": payload.name,
        "email": email,
        "password_hash": hash_password(payload.password),
        "role": "client",
        "serial_number": next_serial,
        "client_id": client_id,
        "referral_code": code,
        "referred_by_code": referred_by,
        "address_line1": "",
        "address_line2": "",
        "bank_name": "",
        "account_holder": "",
        "account_number": "",
        "ifsc": "",
        "kyc_status": "not_started",
        "email_verified": True,
        "mobile_verified": True,
        "created_at": datetime.now(timezone.utc),
    }
    
    res = await db.users.insert_one(doc)
    doc["_id"] = res.inserted_id
    
    print(f"\n[EMAIL OTP] Generated email verification code {otp_code} for user {email}\n")
    logger.info("Sent email verification OTP %s to %s", otp_code, email)
    
    token = create_access_token(str(res.inserted_id), email, "client")
    response.set_cookie("access_token", token, httponly=True, samesite="lax", max_age=ACCESS_TOKEN_MIN * 60, path="/")
    return {"token": token, "user": serialize_user(doc), "dev_otp": otp_code}



@api.post("/auth/login")
async def login(payload: LoginIn, response: Response):
    email = payload.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(payload.password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    token = create_access_token(str(user["_id"]), email, user.get("role", "client"))
    response.set_cookie("access_token", token, httponly=True, samesite="lax", max_age=ACCESS_TOKEN_MIN * 60, path="/")
    return {"token": token, "user": serialize_user(user)}


@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}


@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return serialize_user(user)


@api.post("/auth/send-email-otp")
async def send_email_otp(payload: SendEmailOtpIn):
    email = payload.email.lower()
    user = await db.users.find_one({"email": email})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
        
    otp_code = "".join(secrets.choice(string.digits) for _ in range(6))
    otp_expiry = datetime.now(timezone.utc) + timedelta(minutes=10)
    
    await db.users.update_one(
        {"email": email},
        {"$set": {"email_otp": otp_code, "email_otp_expiry": otp_expiry}}
    )
    
    print(f"\n[EMAIL OTP] Generated email verification code {otp_code} for user {email}\n")
    logger.info("Sent email verification OTP %s to %s", otp_code, email)
    return {"ok": True, "message": "OTP sent successfully (simulated)", "dev_otp": otp_code}


@api.post("/auth/verify-email-otp")
async def verify_email_otp(payload: VerifyEmailOtpIn):
    email = payload.email.lower()
    user = await db.users.find_one({"email": email})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
        
    saved_otp = user.get("email_otp")
    expiry = user.get("email_otp_expiry")
    
    if not saved_otp or saved_otp != payload.otp:
        raise HTTPException(status_code=400, detail="Invalid OTP code")
        
    if isinstance(expiry, str):
        expiry = datetime.fromisoformat(expiry)
    if expiry and expiry.tzinfo is None:
        expiry = expiry.replace(tzinfo=timezone.utc)
        
    if expiry and expiry < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="OTP has expired")
        
    await db.users.update_one(
        {"email": email},
        {"$set": {"email_verified": True}, "$unset": {"email_otp": "", "email_otp_expiry": ""}}
    )
    return {"ok": True, "message": "Email verified successfully"}


@api.post("/auth/send-mobile-otp")
async def send_mobile_otp(payload: SendMobileOtpIn, user: dict = Depends(get_current_user)):
    phone = payload.phone.strip()
    if not phone:
        raise HTTPException(status_code=400, detail="Phone number is required")
        
    otp_code = "".join(secrets.choice(string.digits) for _ in range(6))
    otp_expiry = datetime.now(timezone.utc) + timedelta(minutes=10)
    
    await db.users.update_one(
        {"_id": user["_id"]},
        {"$set": {"mobile_otp": otp_code, "mobile_otp_expiry": otp_expiry, "phone": phone}}
    )
    
    print(f"\n[MOBILE OTP] Generated mobile verification code {otp_code} for user {user.get('email')} ({phone})\n")
    logger.info("Sent mobile verification OTP %s to %s", otp_code, phone)
    return {"ok": True, "message": "OTP sent successfully (simulated)", "dev_otp": otp_code}


@api.post("/auth/verify-mobile-otp")
async def verify_mobile_otp(payload: VerifyMobileOtpIn, user: dict = Depends(get_current_user)):
    phone = payload.phone.strip()
    saved_otp = user.get("mobile_otp")
    expiry = user.get("mobile_otp_expiry")
    
    if not saved_otp or saved_otp != payload.otp:
        raise HTTPException(status_code=400, detail="Invalid OTP code")
        
    if isinstance(expiry, str):
        expiry = datetime.fromisoformat(expiry)
    if expiry and expiry.tzinfo is None:
        expiry = expiry.replace(tzinfo=timezone.utc)
        
    if expiry and expiry < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="OTP has expired")
        
    await db.users.update_one(
        {"_id": user["_id"]},
        {"$set": {"mobile_verified": True, "phone": phone}, "$unset": {"mobile_otp": "", "mobile_otp_expiry": ""}}
    )
    return {"ok": True, "message": "Mobile number verified successfully"}



# -------------------- Client Routes --------------------
@api.put("/users/profile")
async def update_profile(payload: ProfileIn, user: dict = Depends(get_current_user)):
    data = {k: v for k, v in payload.model_dump().items() if v is not None}
    
    # Check for duplicate email updates
    if "email" in data and data["email"].lower() != user["email"].lower():
        email_lower = data["email"].lower()
        existing = await db.users.find_one({"email": email_lower})
        if existing:
            raise HTTPException(status_code=400, detail="Email address is already in use by another account")
        data["email"] = email_lower
        
    await db.users.update_one({"_id": user["_id"]}, {"$set": data})
    updated = await db.users.find_one({"_id": user["_id"]})
    return serialize_user(updated)


@api.put("/users/change-password")
async def change_password(payload: ChangePasswordIn, user: dict = Depends(get_current_user)):
    new_hash = hash_password(payload.new_password)
    await db.users.update_one({"_id": user["_id"]}, {"$set": {"password_hash": new_hash}})
    return {"ok": True, "message": "Password changed successfully"}


@api.post("/deposits")
async def create_deposit(payload: DepositIn, user: dict = Depends(get_current_user)):
    cfg = await db.settings.find_one({"_id": "global_config"})
    min_dep = cfg.get("min_deposit", MIN_DEPOSIT) if cfg else MIN_DEPOSIT

    if user.get("kyc_status") != "verified":
        raise HTTPException(status_code=403, detail="KYC verification required before making a deposit")
    if payload.amount < min_dep:
        raise HTTPException(status_code=400, detail=f"Minimum deposit is ₹{min_dep:,}")
    if not payload.payment_screenshot:
        raise HTTPException(status_code=400, detail="Payment screenshot is required")
        
    now = datetime.now(timezone.utc)
    doc = {
        "user_id": user["_id"],
        "amount": float(payload.amount),
        "payment_method": payload.payment_method,
        "transaction_id": payload.transaction_id,
        "deposit_date": payload.deposit_date,
        "remarks": payload.remarks or "",
        "deposited_at": now,
        "lock_until": None,
        "approved_at": None,
        "status": "pending",
        "payment_screenshot": payload.payment_screenshot,
    }
    res = await db.deposits.insert_one(doc)
    return {
        "id": str(res.inserted_id),
        "amount": doc["amount"],
        "deposited_at": now.isoformat(),
        "lock_until": None,
        "status": "pending",
    }



@api.get("/deposits/my")
async def my_deposits(user: dict = Depends(get_current_user)):
    deposits = await db.deposits.find({"user_id": user["_id"]}).sort("deposited_at", -1).to_list(500)
    out = []
    for d in deposits:
        out.append({
            "id": str(d["_id"]),
            "amount": d["amount"],
            "deposited_at": d["deposited_at"].isoformat() if isinstance(d["deposited_at"], datetime) else d["deposited_at"],
            "lock_until": d["lock_until"].isoformat() if isinstance(d["lock_until"], datetime) and d["lock_until"] else d.get("lock_until"),
            "status": d.get("status", "pending"),
        })
    return out


@api.get("/deposits/lock-status")
async def get_deposit_lock_status(user: dict = Depends(get_current_user)):
    now = datetime.now(timezone.utc)
    approved_deposits = await db.deposits.find({"user_id": user["_id"], "status": "approved"}).sort("approved_at", -1).to_list(500)
    
    cfg = await db.settings.find_one({"_id": "global_config"})
    lock_months = int(cfg.get("lock_period_months", LOCK_MONTHS)) if cfg else LOCK_MONTHS
    
    out = []
    for d in approved_deposits:
        amt = float(d.get("amount", 0))
        dep_date = d.get("deposit_date") or ""
        app_at = d.get("approved_at")
        lu_at = d.get("lock_until")
        
        if isinstance(app_at, str):
            app_at = datetime.fromisoformat(app_at)
        if isinstance(lu_at, str):
            lu_at = datetime.fromisoformat(lu_at)
            
        if app_at and app_at.tzinfo is None:
            app_at = app_at.replace(tzinfo=timezone.utc)
        if lu_at and lu_at.tzinfo is None:
            lu_at = lu_at.replace(tzinfo=timezone.utc)
            
        if app_at and lu_at:
            total_days = (lu_at - app_at).days
            days_remaining = max(0, (lu_at - now).days)
            days_completed = max(0, (now - app_at).days)
            
            months_completed = min(lock_months, int(days_completed // 30))
            months_remaining = max(0, lock_months - months_completed)
            
            progress_percent = min(100, int((days_completed / total_days) * 100)) if total_days > 0 else 100
            
            monthly_checklist = []
            for m in range(1, lock_months + 1):
                m_threshold = app_at + timedelta(days=30 * m)
                if now >= m_threshold:
                    status = "completed"
                elif now >= (app_at + timedelta(days=30 * (m - 1))):
                    status = "active"
                else:
                    status = "pending"
                monthly_checklist.append({"month": m, "status": status})
                
            lock_status = "completed" if days_remaining <= 0 else "locked"
        else:
            total_days = 0
            days_remaining = 0
            days_completed = 0
            months_completed = 0
            months_remaining = lock_months
            progress_percent = 0
            monthly_checklist = [{"month": m, "status": "pending"} for m in range(1, lock_months + 1)]
            lock_status = "locked"
            
        out.append({
            "id": str(d["_id"]),
            "amount": amt,
            "deposit_date": dep_date,
            "approved_at": app_at.isoformat() if app_at else "",
            "lock_until": lu_at.isoformat() if lu_at else "",
            "lock_duration_months": lock_months,
            "months_completed": months_completed,
            "months_remaining": months_remaining,
            "days_completed": days_completed,
            "days_remaining": days_remaining,
            "progress_percent": progress_percent,
            "monthly_checklist": monthly_checklist,
            "lock_status": lock_status,
        })
    return out



@api.get("/profits/my")
async def my_profits(user: dict = Depends(get_current_user)):
    rows = await db.profit_credits.find({"user_id": user["_id"]}).sort("credited_at", -1).to_list(500)
    return [
        {
            "id": str(r["_id"]),
            "amount": r["amount"],
            "note": r.get("note", ""),
            "credited_at": r["credited_at"].isoformat() if isinstance(r["credited_at"], datetime) else r["credited_at"],
        }
        for r in rows
    ]


@api.get("/earnings/summary")
async def earnings_summary(user: dict = Depends(get_current_user)):
    e = await compute_user_balances(user["_id"])
    return e


@api.get("/earnings/daily")
async def earnings_daily(days: int = 30, user: dict = Depends(get_current_user)):
    """Returns daily profit credited by admin for last N days."""
    profits = await db.profit_credits.find({"user_id": user["_id"]}).to_list(5000)
    by_day = {}
    for p in profits:
        d = p["credited_at"]
        if isinstance(d, str):
            d = datetime.fromisoformat(d)
        key = d.date().isoformat()
        by_day[key] = by_day.get(key, 0.0) + float(p["amount"])
    now = datetime.now(timezone.utc)
    return [
        {
            "date": (now - timedelta(days=i)).date().isoformat(),
            "earning": round(by_day.get((now - timedelta(days=i)).date().isoformat(), 0.0), 2),
        }
        for i in range(days - 1, -1, -1)
    ]


@api.get("/referrals/my")
async def my_referrals(user: dict = Depends(get_current_user)):
    """Returns downline tree rooted at current user."""
    tree = await build_referral_tree(user["referral_code"])
    return tree


async def build_referral_tree(referral_code: str, max_depth: int = 1) -> dict:
    node_user = await db.users.find_one({"referral_code": referral_code})
    if not node_user:
        return {}
    children = []
    if max_depth > 0:
        children_cursor = db.users.find({"referred_by_code": referral_code})
        async for c in children_cursor:
            sub = await build_referral_tree(c["referral_code"], max_depth - 1)
            children.append(sub)
    principal = await user_total_deposit(node_user["_id"])
    return {
        "id": str(node_user["_id"]),
        "name": node_user["name"],
        "email": node_user["email"],
        "referral_code": node_user["referral_code"],
        "principal": principal,
        "children": children,
    }



# -------------------- Admin Routes --------------------
@api.get("/admin/stats")
async def admin_stats(admin: dict = Depends(require_admin)):
    users_count = await db.users.count_documents({"role": "client"})
    total_revenue = 0.0
    async for d in db.deposits.find({"status": "approved"}):
        total_revenue += float(d["amount"])
    
    # Total Payout is sum of all approved withdrawals
    total_payout = 0.0
    async for w in db.withdrawals.find({"status": "approved"}):
        total_payout += float(w["amount"])

    now = datetime.now(timezone.utc)
    month_start = datetime(now.year, now.month, 1, tzinfo=timezone.utc)
    
    # Monthly Payout is sum of approved withdrawals in the current month
    monthly_payout = 0.0
    async for w in db.withdrawals.find({"status": "approved", "approved_at": {"$gte": month_start}}):
        monthly_payout += float(w["amount"])

    pending = await db.deposits.count_documents({"status": "pending"})
    pending_wd = await db.withdrawals.count_documents({"status": "pending"})
    return {
        "total_clients": users_count,
        "total_revenue": round(total_revenue, 2),
        "total_payout": round(total_payout, 2),
        "monthly_payout": round(monthly_payout, 2),
        "active_principal": round(total_revenue, 2),
        "pending_deposits": pending,
        "pending_withdrawals": pending_wd,
    }


@api.get("/admin/daily-chart")
async def admin_daily_chart(days: int = 30, admin: dict = Depends(require_admin)):
    deposits = await db.deposits.find({"status": "approved"}).to_list(10000)
    profits = await db.profit_credits.find({}).to_list(10000)
    now = datetime.now(timezone.utc)
    series = []
    for i in range(days - 1, -1, -1):
        day = (now - timedelta(days=i)).date()
        principal = 0.0
        new_deposit = 0.0
        payout = 0.0
        for d in deposits:
            dd = d.get("approved_at") or d["deposited_at"]
            if isinstance(dd, str):
                dd = datetime.fromisoformat(dd)
            if dd.date() <= day:
                principal += float(d["amount"])
            if dd.date() == day:
                new_deposit += float(d["amount"])
        for p in profits:
            dt = p["credited_at"]
            if isinstance(dt, str):
                dt = datetime.fromisoformat(dt)
            if dt.date() == day:
                payout += float(p["amount"])
        series.append({
            "date": day.isoformat(),
            "new_deposit": new_deposit,
            "payout": round(payout, 2),
            "principal": principal,
        })
    return series


@api.get("/admin/deposits/pending")
async def admin_pending_deposits(admin: dict = Depends(require_admin)):
    rows = await db.deposits.find({"status": "pending"}).sort("deposited_at", -1).to_list(1000)
    out = []
    for d in rows:
        u = await db.users.find_one({"_id": d["user_id"]})
        out.append({
            "id": str(d["_id"]),
            "user_id": str(d["user_id"]),
            "user_name": u.get("name") if u else "—",
            "user_email": u.get("email") if u else "—",
            "amount": d["amount"],
            "deposited_at": d["deposited_at"].isoformat() if isinstance(d["deposited_at"], datetime) else d["deposited_at"],
        })
    return out


@api.post("/admin/deposits/{deposit_id}/approve")
async def admin_approve_deposit(deposit_id: str, admin: dict = Depends(require_admin)):
    now = datetime.now(timezone.utc)
    cfg = await db.settings.find_one({"_id": "global_config"})
    lock_months = int(cfg.get("lock_period_months", LOCK_MONTHS)) if cfg else LOCK_MONTHS
    
    res = await db.deposits.update_one(
        {"_id": ObjectId(deposit_id), "status": "pending"},
        {"$set": {"status": "approved", "approved_at": now, "lock_until": now + timedelta(days=30 * lock_months)}},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Pending deposit not found")
    return {"ok": True}



@api.post("/admin/deposits/{deposit_id}/reject")
async def admin_reject_deposit(deposit_id: str, admin: dict = Depends(require_admin)):
    res = await db.deposits.update_one(
        {"_id": ObjectId(deposit_id), "status": "pending"},
        {"$set": {"status": "rejected", "rejected_at": datetime.now(timezone.utc)}},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Pending deposit not found")
    return {"ok": True}


@api.post("/admin/clients/{user_id}/credit-profit")
async def admin_credit_profit(user_id: str, payload: CreditProfitIn, admin: dict = Depends(require_admin)):
    u = await db.users.find_one({"_id": ObjectId(user_id)})
    if not u:
        raise HTTPException(status_code=404, detail="Client not found")
    doc = {
        "user_id": ObjectId(user_id),
        "amount": float(payload.amount),
        "note": payload.note or "",
        "credited_at": datetime.now(timezone.utc),
        "credited_by": admin["email"],
    }
    res = await db.profit_credits.insert_one(doc)
    return {"id": str(res.inserted_id), "ok": True}


@api.post("/admin/reset-history")
async def admin_reset_history(admin: dict = Depends(require_admin)):
    await db.deposits.delete_many({})
    await db.profit_credits.delete_many({})
    await db.withdrawals.delete_many({})
    await db.system_state.delete_many({})
    return {"ok": True}


@api.post("/admin/bulk-credit-profit")
async def admin_bulk_credit(payload: BulkCreditIn, admin: dict = Depends(require_admin)):
    if payload.rate_percent is None and payload.flat_amount is None:
        raise HTTPException(status_code=400, detail="Provide rate_percent or flat_amount")
    
    query = {"role": "client"}
    if payload.client_ids:
        query["_id"] = {"$in": [ObjectId(cid) for cid in payload.client_ids]}
        
    clients = await db.users.find(query).to_list(5000)
    now = datetime.now(timezone.utc)
    count = 0
    total = 0.0
    for c in clients:
        principal = await user_total_deposit(c["_id"])
        if principal <= 0 and payload.rate_percent is not None:
            continue
        amt = float(payload.flat_amount) if payload.flat_amount else round(principal * (payload.rate_percent / 100.0), 2)
        if amt <= 0:
            continue
        await db.profit_credits.insert_one({
            "user_id": c["_id"],
            "amount": amt,
            "note": payload.note or "Bulk credit",
            "credited_at": now,
            "credited_by": admin["email"],
        })
        count += 1
        total += amt
    return {"credited_clients": count, "total_credited": round(total, 2)}


@api.post("/admin/clients/{user_id}/kyc")
async def admin_set_kyc(user_id: str, payload: KycStatusIn, admin: dict = Depends(require_admin)):
    if payload.status not in ("pending", "verified", "rejected"):
        raise HTTPException(status_code=400, detail="Invalid status")
    res = await db.users.update_one({"_id": ObjectId(user_id), "role": "client"}, {"$set": {"kyc_status": payload.status}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    return {"ok": True, "kyc_status": payload.status}


@api.post("/withdrawals")
async def create_withdrawal(payload: WithdrawIn, user: dict = Depends(get_current_user)):
    balances = await compute_user_balances(user["_id"])
    withdrawable = balances["withdrawable_amount"]
    if payload.amount > withdrawable:
        raise HTTPException(status_code=400, detail=f"Only ₹{withdrawable:,.2f} available for withdrawal")
    
    now = datetime.now(timezone.utc)
    doc = {
        "user_id": user["_id"],
        "amount": float(payload.amount),
        "status": "pending",
        "requested_at": now,
        "bank_name": user.get("bank_name", ""),
        "account_holder": user.get("account_holder", ""),
        "account_number": user.get("account_number", ""),
        "ifsc": user.get("ifsc", ""),
    }
    res = await db.withdrawals.insert_one(doc)
    return {"id": str(res.inserted_id), "amount": doc["amount"], "status": "pending"}


class UpdateWithdrawalStatusIn(BaseModel):
    status: str  # "pending" | "under_review" | "approved" | "rejected" | "paid"
    transaction_id: Optional[str] = ""


@api.get("/withdrawals/my")
async def my_withdrawals(user: dict = Depends(get_current_user)):
    rows = await db.withdrawals.find({"user_id": user["_id"]}).sort("requested_at", -1).to_list(500)
    return [
        {
            "id": str(r["_id"]),
            "amount": r["amount"],
            "status": r["status"],
            "transaction_id": r.get("transaction_id", ""),
            "requested_at": r["requested_at"].isoformat() if isinstance(r["requested_at"], datetime) else r["requested_at"],
            "bank_name": r.get("bank_name", ""),
            "account_number": r.get("account_number", ""),
            "ifsc": r.get("ifsc", ""),
        }
        for r in rows
    ]


@api.get("/admin/withdrawals/pending")
async def admin_pending_withdrawals(admin: dict = Depends(require_admin)):
    # Returns both pending and under_review withdrawals for handling
    rows = await db.withdrawals.find({"status": {"$in": ["pending", "under_review", "approved"]}}).sort("requested_at", -1).to_list(1000)
    out = []
    for r in rows:
        u = await db.users.find_one({"_id": r["user_id"]})
        out.append({
            "id": str(r["_id"]),
            "user_name": u.get("name") if u else "—",
            "user_email": u.get("email") if u else "—",
            "amount": r["amount"],
            "status": r["status"],
            "transaction_id": r.get("transaction_id", ""),
            "requested_at": r["requested_at"].isoformat() if isinstance(r["requested_at"], datetime) else r["requested_at"]
        })
    return out


@api.post("/admin/withdrawals/{wid}/status")
async def admin_update_withdrawal_status(wid: str, payload: UpdateWithdrawalStatusIn, admin: dict = Depends(require_admin)):
    status = payload.status.lower().strip()
    if status not in ["pending", "under_review", "approved", "rejected", "paid"]:
        raise HTTPException(status_code=400, detail="Invalid status value")
        
    update_fields = {"status": status}
    if payload.transaction_id:
        update_fields["transaction_id"] = payload.transaction_id.strip()
        
    if status in ["approved", "paid"]:
        update_fields["approved_at"] = datetime.now(timezone.utc)
        
    res = await db.withdrawals.update_one({"_id": ObjectId(wid)}, {"$set": update_fields})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Withdrawal request not found")
    return {"ok": True}


@api.post("/admin/withdrawals/{wid}/approve")
async def admin_approve_withdrawal(wid: str, admin: dict = Depends(require_admin)):
    res = await db.withdrawals.update_one({"_id": ObjectId(wid)}, {"$set": {"status": "approved", "approved_at": datetime.now(timezone.utc)}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Withdrawal request not found")
    return {"ok": True}


@api.post("/admin/withdrawals/{wid}/reject")
async def admin_reject_withdrawal(wid: str, admin: dict = Depends(require_admin)):
    res = await db.withdrawals.update_one({"_id": ObjectId(wid)}, {"$set": {"status": "rejected"}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Withdrawal request not found")
    return {"ok": True}



@api.get("/admin/clients/export")
async def admin_clients_export(admin: dict = Depends(require_admin)):
    users = await db.users.find({"role": "client"}).sort("created_at", -1).to_list(2000)
    rows = []
    for u in users:
        earn = await compute_user_balances(u["_id"])
        rows.append({
            "Name": u.get("name"),
            "Email": u.get("email"),
            "Referral Code": u.get("referral_code"),
            "Referred By": u.get("referred_by_code") or "",
            "Address Line 1": u.get("address_line1", ""),
            "Address Line 2": u.get("address_line2", ""),
            "Bank Name": u.get("bank_name", ""),
            "Account Holder": u.get("account_holder", ""),
            "Account Number": u.get("account_number", ""),
            "IFSC": u.get("ifsc", ""),
            "Nominee": u.get("nominee", ""),

            "Principal (INR)": earn["principal"],
            "Total Earned (INR)": earn["total_earned"],
            "Daily Earning (INR)": earn["today_earning"],
            "Joined": u.get("created_at").isoformat() if u.get("created_at") else "",
        })
    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Clients", index=False)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="kanak_clients.xlsx"'},
    )


@api.get("/admin/clients/{user_id}")
async def admin_client_detail(user_id: str, admin: dict = Depends(require_admin)):
    u = await db.users.find_one({"_id": ObjectId(user_id)})
    if not u:
        raise HTTPException(status_code=404, detail="Client not found")
    earn = await compute_user_balances(u["_id"])
    deposits = await db.deposits.find({"user_id": u["_id"]}).sort("deposited_at", -1).to_list(500)
    profits = await db.profit_credits.find({"user_id": u["_id"]}).sort("credited_at", -1).to_list(500)
    # Include referral credits in profits view or separate them
    referrals = await db.referral_credits.find({"user_id": u["_id"]}).sort("credited_at", -1).to_list(500)

    return {
        "user": {**serialize_user(u), **earn},
        "deposits": [
            {
                "id": str(d["_id"]),
                "amount": d["amount"],
                "status": d.get("status", "pending"),
                "deposited_at": d["deposited_at"].isoformat() if isinstance(d["deposited_at"], datetime) else d["deposited_at"],
                "lock_until": d["lock_until"].isoformat() if isinstance(d["lock_until"], datetime) else d["lock_until"],
            }
            for d in deposits
        ],
        "profits": [
            {
                "id": str(p["_id"]),
                "amount": p["amount"],
                "note": p.get("note", ""),
                "credited_at": p["credited_at"].isoformat() if isinstance(p["credited_at"], datetime) else p["credited_at"],
            }
            for p in profits
        ],
    }


@api.put("/admin/clients/{user_id}")
async def admin_edit_client(user_id: str, payload: AdminEditClientIn, admin: dict = Depends(require_admin)):
    data = {k: v for k, v in payload.model_dump().items() if v is not None}
    if "email" in data:
        data["email"] = data["email"].lower()
    if not data:
        raise HTTPException(status_code=400, detail="No changes")
    res = await db.users.update_one({"_id": ObjectId(user_id), "role": "client"}, {"$set": data})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    updated = await db.users.find_one({"_id": ObjectId(user_id)})
    return serialize_user(updated)


@api.delete("/admin/clients/{user_id}")
async def admin_delete_client(user_id: str, admin: dict = Depends(require_admin)):
    oid = ObjectId(user_id)
    u = await db.users.find_one({"_id": oid, "role": "client"})
    if not u:
        raise HTTPException(status_code=404, detail="Client not found")
    await db.deposits.delete_many({"user_id": oid})
    await db.profit_credits.delete_many({"user_id": oid})
    await db.users.delete_one({"_id": oid})
    return {"ok": True}


@api.get("/admin/clients")
async def admin_clients(admin: dict = Depends(require_admin)):
    users = await db.users.find({"role": "client"}).sort("created_at", -1).to_list(2000)
    out = []
    for u in users:
        earn = await compute_user_balances(u["_id"])
        item = serialize_user(u)
        item["principal"] = earn["principal"]
        item["total_earned"] = earn["total_earned"]
        item["today_earning"] = earn["today_earning"]
        item["kyc_status"] = u.get("kyc_status", "not_started")
        item["referral_count"] = earn["referral_count"]
        out.append(item)
    return out


# Route moved above detail route to resolve routing collision


@api.get("/admin/reports/monthly-payout")
async def download_monthly_payout_report(
    month: int,
    year: int,
    client_name: Optional[str] = None,
    payment_status: Optional[str] = "All",
    client_ids: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    import io
    import tempfile
    import os
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from PIL import Image as PILImage, ImageDraw

    # 1. Fetch filtered clients
    query = {"role": "client"}
    if client_ids:
        ids = [ObjectId(x) for x in client_ids.split(",") if x.strip()]
        query["_id"] = {"$in": ids}

    clients = await db.users.find(query).to_list(10000)

    if client_name:
        name_q = client_name.strip().lower()
        clients = [c for c in clients if name_q in c.get("name", "").lower()]

    # Time frame limits
    month_start = datetime(year, month, 1, tzinfo=timezone.utc)
    if month == 12:
        month_end = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
    else:
        month_end = datetime(year, month + 1, 1, tzinfo=timezone.utc)

    month_name = month_start.strftime("%B %Y")
    
    # Fetch settings for rate
    cfg = await db.settings.find_one({"_id": "global_config"})
    profit_pct_daily = cfg.get("profit_percentage_daily", 0.10) if cfg else 0.10
    monthly_profit_rate = profit_pct_daily * 30.0

    sheet1_rows = []
    sheet2_rows = []

    for c in clients:
        client_id = c.get("client_id", "—")
        name = c.get("name", "—")
        email = c.get("email", "—")
        phone = c.get("phone") or c.get("company_sms") or c.get("company_whatsapp") or "—"
        kyc_status = c.get("kyc_status", "not_started")

        # Balances
        balances = await compute_user_balances(c["_id"])
        principal = balances["principal"]
        wallet_balance = balances["withdrawable_amount"]

        # Referrals stats
        referred_users = await db.users.find({"referred_by_code": c.get("referral_code")}).to_list(10000)
        total_referrals_count = len(referred_users)

        active_referrals_count = 0
        inactive_referrals_count = 0
        total_investment_by_referrals = 0.0

        for ru in referred_users:
            ru_bal = await compute_user_balances(ru["_id"])
            ru_principal = ru_bal["principal"]
            if ru_principal > 0:
                active_referrals_count += 1
            else:
                inactive_referrals_count += 1
            total_investment_by_referrals += ru_principal

        # Monthly Profit calculations
        monthly_profit_amount = 0.0
        async for p in db.profit_credits.find({"user_id": c["_id"], "credited_at": {"$gte": month_start, "$lt": month_end}}):
            monthly_profit_amount += float(p["amount"])

        # Payout status based on withdrawals requested this month
        wdr_list = await db.withdrawals.find({"user_id": c["_id"], "requested_at": {"$gte": month_start, "$lt": month_end}}).to_list(100)
        has_pending_wdr = any(w["status"] == "pending" for w in wdr_list)
        approved_wdrs = [w for w in wdr_list if w["status"] == "approved"]

        if has_pending_wdr:
            payout_status = "Pending"
            payout_date = "—"
        elif approved_wdrs:
            payout_status = "Paid"
            app_at = approved_wdrs[-1].get("approved_at")
            payout_date = app_at.strftime("%Y-%m-%d") if app_at else approved_wdrs[-1].get("requested_at").strftime("%Y-%m-%d")
        else:
            payout_status = "Paid" if monthly_profit_amount > 0 else "—"
            payout_date = month_end.strftime("%Y-%m-%d") if monthly_profit_amount > 0 else "—"

        # Apply payout status filter
        if payment_status == "All" or payout_status == payment_status:
            sheet1_rows.append({
                "client_id": client_id,
                "name": name,
                "email": email,
                "phone": phone,
                "principal": principal,
                "profit_rate": monthly_profit_rate,
                "profit_amount": monthly_profit_amount,
                "month": month_name,
                "status": payout_status,
                "payment_date": payout_date,
                "wallet_balance": wallet_balance,
                "kyc_status": kyc_status,
                "referrals_count": total_referrals_count,
                "active_referrals": active_referrals_count,
                "inactive_referrals": inactive_referrals_count,
                "referrals_investment": total_investment_by_referrals
            })

        # Referral payout calculations
        direct_referral_income = 0.0
        async for r in db.referral_credits.find({"user_id": c["_id"], "credited_at": {"$gte": month_start, "$lt": month_end}}):
            direct_referral_income += float(r["amount"])

        level_income = 0.0
        total_referral_commission = direct_referral_income + level_income

        if has_pending_wdr:
            ref_status = "Pending"
            ref_date = "—"
        elif approved_wdrs:
            ref_status = "Paid"
            app_at = approved_wdrs[-1].get("approved_at")
            ref_date = app_at.strftime("%Y-%m-%d") if app_at else approved_wdrs[-1].get("requested_at").strftime("%Y-%m-%d")
        else:
            ref_status = "Paid" if total_referral_commission > 0 else "—"
            ref_date = month_end.strftime("%Y-%m-%d") if total_referral_commission > 0 else "—"

        # Apply referral status filter
        if payment_status == "All" or ref_status == payment_status:
            sheet2_rows.append({
                "client_id": client_id,
                "name": name,
                "email": email,
                "total_referrals": total_referrals_count,
                "direct_income": direct_referral_income,
                "level_income": level_income,
                "total_commission": total_referral_commission,
                "month": month_name,
                "status": ref_status,
                "payment_date": ref_date,
                "referrals_count": total_referrals_count,
                "active_referrals": active_referrals_count,
                "inactive_referrals": inactive_referrals_count,
                "referrals_investment": total_investment_by_referrals
            })

    # 2. Build workbook
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Monthly Profit Payout"
    ws2 = wb.create_sheet(title="Referral Payout")

    # Font and styling configs
    TITLE_FONT = Font(name="Segoe UI", size=14, bold=True, color="002FA7")
    SUBTITLE_FONT = Font(name="Segoe UI", size=9, italic=True, color="475569")
    HEADER_FILL = PatternFill(start_color="002FA7", end_color="002FA7", fill_type="solid")
    HEADER_FONT = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    DATA_FONT = Font(name="Segoe UI", size=10, color="000000")
    BOLD_FONT = Font(name="Segoe UI", size=10, bold=True, color="000000")
    ALT_ROW_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    TOTAL_ROW_FILL = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    
    THIN_BORDER = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    TOTAL_BORDER = Border(
        top=Side(style='thin', color='94A3B8'),
        bottom=Side(style='double', color='000000'),
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0')
    )

    logo_path = os.path.join(tempfile.gettempdir(), "kanak_logo.png")
    logo_exists = False
    try:
        img = PILImage.new('RGB', (240, 60), color='#002FA7')
        d = ImageDraw.Draw(img)
        d.rectangle([(5, 5), (50, 55)], fill='#FFFFFF')
        d.text((23, 18), "K", fill="#002FA7")
        d.text((65, 22), "KANAK INFOSYS", fill="#FFFFFF")
        img.save(logo_path)
        logo_exists = True
    except Exception:
        pass

    for ws, title_text, rows_list, headers in [
        (ws1, "MONTHLY PROFIT PAYOUT REPORT", sheet1_rows, [
            "Client ID", "Client Name", "Email", "Mobile Number", "Principal Investment", 
            "Monthly Profit Rate (%)", "Monthly Profit Amount", "Current Month", 
            "Payment Status", "Payment Date", "Wallet Balance", "KYC Status",
            "Number of Referrals", "Active Referrals", "Inactive Referrals", "Total Investment by Referrals"
        ]),
        (ws2, "REFERRAL PAYOUT REPORT", sheet2_rows, [
            "Client ID", "Client Name", "Email", "Total Referrals", "Direct Referral Income", 
            "Level Income", "Total Referral Commission", "Current Month", 
            "Payment Status", "Payment Date", 
            "Number of Referrals", "Active Referrals", "Inactive Referrals", "Total Investment by Referrals"
        ])
    ]:
        ws.row_dimensions[2].height = 45
        if logo_exists:
            img_obj = OpenpyxlImage(logo_path)
            ws.add_image(img_obj, 'A2')

        ws.merge_cells('D2:P2')
        ws['D2'] = title_text
        ws['D2'].font = TITLE_FONT
        ws['D2'].alignment = Alignment(vertical="center")

        ws.merge_cells('D3:P3')
        ws['D3'] = f"Month: {month_name} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Status: {payment_status}"
        ws['D3'].font = SUBTITLE_FONT
        ws['D3'].alignment = Alignment(vertical="center")

        ws.append([]) # row 1
        ws.append([]) # row 2
        ws.append([]) # row 3
        ws.append([]) # row 4
        ws.append(headers) # row 5

        # Format header row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=5, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

        # Data rows
        start_row = 6
        for idx, r in enumerate(rows_list, start=start_row):
            if ws == ws1:
                row_data = [
                    r["client_id"], r["name"], r["email"], r["phone"], r["principal"],
                    r["profit_rate"] / 100.0, r["profit_amount"], r["month"],
                    r["status"], r["payment_date"], r["wallet_balance"], r["kyc_status"],
                    r["referrals_count"], r["active_referrals"], r["inactive_referrals"], r["referrals_investment"]
                ]
            else:
                row_data = [
                    r["client_id"], r["name"], r["email"], r["total_referrals"], r["direct_income"],
                    r["level_income"], r["total_commission"], r["month"],
                    r["status"], r["payment_date"],
                    r["referrals_count"], r["active_referrals"], r["inactive_referrals"], r["referrals_investment"]
                ]
            ws.append(row_data)

            is_alt = (idx % 2 == 1)
            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=idx, column=col_idx)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                if is_alt:
                    cell.fill = ALT_ROW_FILL

                # Custom alignments/formats
                if ws == ws1:
                    if col_idx in [5, 7, 11, 16]:
                        cell.number_format = '[$₹-4009] #,##0.00'
                        cell.alignment = Alignment(horizontal="right")
                    elif col_idx == 6:
                        cell.number_format = '0.00%'
                        cell.alignment = Alignment(horizontal="right")
                    elif col_idx in [12, 13, 14, 15]:
                        cell.alignment = Alignment(horizontal="center")
                    elif col_idx in [1, 8, 9, 10]:
                        cell.alignment = Alignment(horizontal="center")
                else:
                    if col_idx in [5, 6, 7, 14]:
                        cell.number_format = '[$₹-4009] #,##0.00'
                        cell.alignment = Alignment(horizontal="right")
                    elif col_idx in [4, 11, 12, 13]:
                        cell.alignment = Alignment(horizontal="center")
                    elif col_idx in [1, 8, 9, 10]:
                        cell.alignment = Alignment(horizontal="center")

        # Total Rows
        end_row = start_row + len(rows_list) - 1
        total_row_idx = end_row + 1
        
        # Avoid empty borders if no records
        if len(rows_list) == 0:
            end_row = 5
            total_row_idx = 6

        ws.row_dimensions[total_row_idx].height = 24
        
        # Prepare totals cells
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=total_row_idx, column=col_idx)
            cell.fill = TOTAL_ROW_FILL
            cell.font = BOLD_FONT
            cell.border = TOTAL_BORDER

        if len(rows_list) > 0:
            if ws == ws1:
                ws.cell(row=total_row_idx, column=1, value="Total Clients")
                ws.cell(row=total_row_idx, column=2, value=f"=COUNTA(A6:A{end_row})")
                
                ws.cell(row=total_row_idx, column=5, value="Total Profit Paid")
                ws.cell(row=total_row_idx, column=7, value=f'=SUMIF(I6:I{end_row}, "Paid", G6:G{end_row})')
                ws.cell(row=total_row_idx, column=7).number_format = '[$₹-4009] #,##0.00'
                
                ws.cell(row=total_row_idx, column=9, value="Total Pending Profit")
                ws.cell(row=total_row_idx, column=11, value=f'=SUMIF(I6:I{end_row}, "Pending", G6:G{end_row})')
                ws.cell(row=total_row_idx, column=11).number_format = '[$₹-4009] #,##0.00'
            else:
                ws.cell(row=total_row_idx, column=1, value="Total Referrals")
                ws.cell(row=total_row_idx, column=4, value=f"=SUM(D6:D{end_row})")
                
                ws.cell(row=total_row_idx, column=5, value="Total Comm Paid")
                ws.cell(row=total_row_idx, column=7, value=f'=SUMIF(I6:I{end_row}, "Paid", G6:G{end_row})')
                ws.cell(row=total_row_idx, column=7).number_format = '[$₹-4009] #,##0.00'
                
                ws.cell(row=total_row_idx, column=9, value="Total Pending Comm")
                ws.cell(row=total_row_idx, column=11, value=f'=SUMIF(I6:I{end_row}, "Pending", G6:G{end_row})')
                ws.cell(row=total_row_idx, column=11).number_format = '[$₹-4009] #,##0.00'

        # Freeze Panes
        ws.freeze_panes = 'A6'

        # Landscape layout
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        # Width fit
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row > 4 and cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    if logo_exists and os.path.exists(logo_path):
        try:
            os.remove(logo_path)
        except Exception:
            pass

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="kanak_monthly_report_{year}_{month}.xlsx"'},
    )


@api.get("/admin/referral-tree")
async def admin_referral_tree(admin: dict = Depends(require_admin)):
    roots = await db.users.find({"role": "client", "referred_by_code": None}).to_list(2000)
    forest = []
    for r in roots:
        forest.append(await build_referral_tree(r["referral_code"]))
    return forest


# -------------------- Client Extensions --------------------
@api.get("/referrals/summary")
async def referrals_summary(user: dict = Depends(get_current_user)):
    directs = await db.users.find({"referred_by_code": user["referral_code"]}).to_list(1000)
        
    direct_list = []
    total_direct_principal = 0.0
    for d in directs:
        p = await user_total_deposit(d["_id"])
        total_direct_principal += p
        direct_list.append({
            "id": str(d["_id"]),
            "name": d["name"],
            "email": d["email"],
            "principal": p,
            "kyc_status": d.get("kyc_status", "not_started"),
            "joined_at": d["created_at"].isoformat() if d.get("created_at") else None,
        })
        
    credits = await db.referral_credits.find({"user_id": user["_id"]}).sort("credited_at", -1).to_list(1000)
    total_earned = sum(float(cr["amount"]) for cr in credits)
    
    cfg = await db.settings.find_one({"_id": "global_config"})
    ref_monthly_pct = cfg.get("referral_percentage_monthly", 1.0) if cfg else 1.0
    monthly_est = total_direct_principal * (ref_monthly_pct / 100.0)
    
    history = []
    for cr in credits:
        ref_u = await db.users.find_one({"_id": cr.get("referred_user_id")})
        history.append({
            "id": str(cr["_id"]),
            "amount": cr["amount"],
            "referred_user_name": ref_u.get("name") if ref_u else "Direct Referral",
            "note": cr.get("note", ""),
            "credited_at": cr["credited_at"].isoformat() if isinstance(cr["credited_at"], datetime) else cr["credited_at"],
        })
        
    return {
        "referral_code": user["referral_code"],
        "total_earnings": round(total_earned, 2),
        "monthly_commission_estimate": round(monthly_est, 2),
        "direct_count": len(directs),
        "indirect_count": 0,
        "direct_referrals": direct_list,
        "indirect_referrals": [],
        "history": history
    }



@api.get("/transactions")
async def get_transactions(user: dict = Depends(get_current_user)):
    deposits = await db.deposits.find({"user_id": user["_id"]}).to_list(500)
    withdrawals = await db.withdrawals.find({"user_id": user["_id"]}).to_list(500)
    profits = await db.profit_credits.find({"user_id": user["_id"]}).to_list(1000)
    referrals = await db.referral_credits.find({"user_id": user["_id"]}).to_list(1000)
    
    txs = []
    
    for d in deposits:
        txs.append({
            "id": f"DEP-{str(d['_id'])}",
            "type": "deposit",
            "amount": d["amount"],
            "status": d.get("status", "pending"),
            "date": d["deposited_at"].isoformat() if isinstance(d["deposited_at"], datetime) else d["deposited_at"],
            "note": "Deposit Request" + (" (Locked for 6 Months)" if d.get("status") == "approved" else "")
        })
        
    for w in withdrawals:
        txs.append({
            "id": f"WDR-{str(w['_id'])}",
            "type": "withdrawal",
            "amount": w["amount"],
            "status": w.get("status", "pending"),
            "date": w["requested_at"].isoformat() if isinstance(w["requested_at"], datetime) else w["requested_at"],
            "note": "Withdrawal Request"
        })
        
    for p in profits:
        txs.append({
            "id": f"PRF-{str(p['_id'])}",
            "type": "profit",
            "amount": p["amount"],
            "status": "approved",
            "date": p["credited_at"].isoformat() if isinstance(p["credited_at"], datetime) else p["credited_at"],
            "note": p.get("note", "Daily Profit Credit")
        })
        
    for r in referrals:
        txs.append({
            "id": f"REF-{str(r['_id'])}",
            "type": "referral",
            "amount": r["amount"],
            "status": "approved",
            "date": r["credited_at"].isoformat() if isinstance(r["credited_at"], datetime) else r["credited_at"],
            "note": r.get("note", "Referral Commission")
        })
        
    txs.sort(key=lambda x: x["date"], reverse=True)
    return txs


# -------------------- Broadcast & Notification Routes --------------------
@api.post("/admin/broadcasts")
async def create_broadcast(payload: BroadcastIn, admin: dict = Depends(require_admin)):
    now = datetime.now(timezone.utc)
    doc = {
        "title": payload.title,
        "description": payload.description,
        "created_at": now
    }
    res = await db.broadcasts.insert_one(doc)
    return {"id": str(res.inserted_id), "ok": True}


@api.put("/admin/broadcasts/{bid}")
async def update_broadcast(bid: str, payload: BroadcastIn, admin: dict = Depends(require_admin)):
    res = await db.broadcasts.update_one(
        {"_id": ObjectId(bid)},
        {"$set": {"title": payload.title, "description": payload.description}}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Broadcast not found")
    return {"ok": True}


@api.delete("/admin/broadcasts/{bid}")
async def delete_broadcast(bid: str, admin: dict = Depends(require_admin)):
    res = await db.broadcasts.delete_one({"_id": ObjectId(bid)})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Broadcast not found")
    await db.user_notifications.delete_many({"broadcast_id": ObjectId(bid)})
    return {"ok": True}


@api.get("/admin/broadcasts")
async def get_admin_broadcasts(admin: dict = Depends(require_admin)):
    rows = await db.broadcasts.find({}).sort("created_at", -1).to_list(1000)
    return [
        {
            "id": str(r["_id"]),
            "title": r["title"],
            "description": r["description"],
            "created_at": r["created_at"].isoformat() if isinstance(r["created_at"], datetime) else r["created_at"]
        }
        for r in rows
    ]


@api.get("/notifications")
async def get_client_notifications(user: dict = Depends(get_current_user)):
    broadcasts = await db.broadcasts.find({}).sort("created_at", -1).to_list(1000)
    user_states = await db.user_notifications.find({"user_id": user["_id"]}).to_list(10000)
    
    state_map = {str(s["broadcast_id"]): s for s in user_states}
    
    out = []
    for b in broadcasts:
        bid_str = str(b["_id"])
        state = state_map.get(bid_str)
        if state and state.get("deleted"):
            continue
        out.append({
            "id": bid_str,
            "title": b["title"],
            "description": b["description"],
            "created_at": b["created_at"].isoformat() if isinstance(b["created_at"], datetime) else b["created_at"],
            "read": state.get("read", False) if state else False
        })
    return out


@api.post("/notifications/{bid}/read")
async def mark_notification_read(bid: str, user: dict = Depends(get_current_user)):
    await db.user_notifications.update_one(
        {"user_id": user["_id"], "broadcast_id": ObjectId(bid)},
        {"$set": {"read": True}},
        upsert=True
    )
    return {"ok": True}


@api.delete("/notifications/{bid}")
async def mark_notification_deleted(bid: str, user: dict = Depends(get_current_user)):
    await db.user_notifications.update_one(
        {"user_id": user["_id"], "broadcast_id": ObjectId(bid)},
        {"$set": {"deleted": True}},
        upsert=True
    )
    return {"ok": True}


# -------------------- Settings Routes --------------------
@api.get("/settings")
async def get_settings():
    cfg = await db.settings.find_one({"_id": "global_config"})
    if not cfg:
        return {
            "min_deposit": 100000.0,
            "profit_percentage_daily": 0.10,
            "referral_percentage_monthly": 1.0,
            "lock_period_months": 6,
            "company_bank_name": "Kanak Infosys Bank",
            "company_account_number": "9876543210123",
            "company_account_holder": "Kanak Infosys Ltd",
            "company_ifsc": "KNK000123",
            "company_email": "support@kanakinfosys.com",
            "company_whatsapp": "+919876543210",
            "company_sms": "+919876543210",
            "maintenance_mode": False,
        }
    return cfg



@api.put("/settings")
async def update_settings(payload: SettingsIn, admin: dict = Depends(require_admin)):
    data = {k: v for k, v in payload.model_dump().items() if v is not None}
    await db.settings.update_one({"_id": "global_config"}, {"$set": data}, upsert=True)
    updated = await db.settings.find_one({"_id": "global_config"})
    return updated


# -------------------- App Bootstrap --------------------
@api.get("/")
async def root():
    return {"service": "kanak-infosys", "status": "ok"}


app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)
app.add_middleware(GZipMiddleware, minimum_size=1000)


@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.users.create_index("referral_code", unique=True)
    await db.users.create_index("referred_by_code")
    await db.deposits.create_index("user_id")
    await db.deposits.create_index("status")
    
    # Indexes to optimize sub-queries and dashboard loads
    await db.profit_credits.create_index("user_id")
    await db.referral_credits.create_index("user_id")
    await db.withdrawals.create_index("user_id")
    await db.withdrawals.create_index("status")
    await db.notifications.create_index("user_id")

    # Seed default global configuration
    existing_cfg = await db.settings.find_one({"_id": "global_config"})
    if not existing_cfg:
        await db.settings.insert_one({
            "_id": "global_config",
            "min_deposit": 100000.0,
            "profit_percentage_daily": 0.10,
            "referral_percentage_monthly": 1.0,
            "lock_period_months": 6,
            "company_bank_name": "Kanak Infosys Bank",
            "company_account_number": "9876543210123",
            "company_account_holder": "Kanak Infosys Ltd",
            "company_ifsc": "KNK000123",
            "company_branch": "Main Branch",
            "company_email": "support@kanakinfosys.com",
            "company_whatsapp": "+919876543210",
            "company_sms": "+919876543210",
            "maintenance_mode": False,
        })
        logger.info("Seeded global settings configuration")
    else:
        if "company_branch" not in existing_cfg:
            await db.settings.update_one({"_id": "global_config"}, {"$set": {"company_branch": "Main Branch"}})

    # Self-healing migration for existing clients
    existing_clients = await db.users.find({"role": "client"}).sort("created_at", 1).to_list(None)
    for idx, c in enumerate(existing_clients, start=50):
        updates = {}
        if "serial_number" not in c:
            updates["serial_number"] = idx
        serial = c.get("serial_number", idx)
        if "client_id" not in c:
            updates["client_id"] = f"CLI{serial:06d}"
        
        # Format code check
        code_ok = False
        ref_code = c.get("referral_code", "")
        if ref_code.startswith("KNK") and len(ref_code) == 7:
            try:
                int(ref_code[3:])
                code_ok = True
            except ValueError:
                pass
        
        if not code_ok:
            new_code = f"KNK{serial:04d}"
            old_code = c.get("referral_code")
            updates["referral_code"] = new_code
            if old_code:
                await db.users.update_many({"referred_by_code": old_code}, {"$set": {"referred_by_code": new_code}})
                
        if updates:
            await db.users.update_one({"_id": c["_id"]}, {"$set": updates})

    admin_email = os.environ["ADMIN_EMAIL"].lower()
    admin_pw = os.environ["ADMIN_PASSWORD"]
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        code = "KNK000"
        await db.users.insert_one({
            "name": "Kanak Admin",
            "email": admin_email,
            "password_hash": hash_password(admin_pw),
            "role": "admin",
            "serial_number": 0,
            "client_id": "CLI000000",
            "referral_code": code,
            "referred_by_code": None,
            "address_line1": "",
            "address_line2": "",
            "bank_name": "",
            "account_holder": "",
            "account_number": "",
            "ifsc": "",
            "email_verified": True,
            "mobile_verified": True,
            "created_at": datetime.now(timezone.utc),
        })
        logger.info("Seeded admin user %s", admin_email)
    else:
        updates = {}
        if not verify_password(admin_pw, existing.get("password_hash", "")):
            updates["password_hash"] = hash_password(admin_pw)
        if "serial_number" not in existing:
            updates["serial_number"] = 0
            updates["client_id"] = "CLI000000"
        if not existing.get("email_verified"):
            updates["email_verified"] = True
        if not existing.get("mobile_verified"):
            updates["mobile_verified"] = True
        if updates:
            await db.users.update_one({"email": admin_email}, {"$set": updates})
            logger.info("Updated admin config for %s", admin_email)


    asyncio.create_task(daily_auto_credit_loop())


async def daily_auto_credit_loop():
    """Every 24h, auto-credit daily profit (0.10%) and referral commission (1% monthly) to approved clients."""
    while True:
        try:
            today = datetime.now(timezone.utc).date().isoformat()
            marker = await db.system_state.find_one({"_id": "last_auto_credit"})
            if not marker or marker.get("date") != today:
                cfg = await db.settings.find_one({"_id": "global_config"})
                daily_profit_rate = (cfg.get("profit_percentage_daily", 0.10) / 100.0) if cfg else (0.10 / 100.0)
                ref_monthly_pct = cfg.get("referral_percentage_monthly", 1.0) if cfg else 1.0
                daily_referral_rate = (ref_monthly_pct / 100.0) / 30.0

                clients = await db.users.find({"role": "client"}).to_list(5000)
                now = datetime.now(timezone.utc)
                total_profit = 0.0
                total_ref = 0.0
                profit_count = 0
                ref_count = 0
                
                for c in clients:
                    principal = await user_total_deposit(c["_id"])
                    if principal <= 0:
                        continue
                    
                    # 1. Daily profit credit
                    amt_profit = round(principal * daily_profit_rate, 2)
                    if amt_profit > 0:
                        exists = await db.profit_credits.find_one({
                            "user_id": c["_id"],
                            "credited_at": {"$gte": datetime.combine(now.date(), datetime.min.time()).replace(tzinfo=timezone.utc)},
                            "note": {"$regex": "daily credit", "$options": "i"}
                        })
                        if not exists:
                            await db.profit_credits.insert_one({
                                "user_id": c["_id"],
                                "amount": amt_profit,
                                "note": f"Automated daily credit ({(daily_profit_rate * 100.0):.2f}%)",
                                "credited_at": now,
                                "credited_by": "system",
                            })
                            total_profit += amt_profit
                            profit_count += 1
                            
                    # 2. Referral Commission credit
                    if c.get("referred_by_code"):
                        created_at = c.get("created_at")
                        if created_at:
                            if isinstance(created_at, str):
                                created_at = datetime.fromisoformat(created_at)
                            if created_at.tzinfo is None:
                                created_at = created_at.replace(tzinfo=timezone.utc)
                            if now - created_at <= timedelta(days=365):
                                referrer = await db.users.find_one({"referral_code": c["referred_by_code"].strip().upper()})
                                if referrer:
                                    amt_ref = round(principal * daily_referral_rate, 2)
                                    if amt_ref > 0:
                                        exists_ref = await db.referral_credits.find_one({
                                            "user_id": referrer["_id"],
                                            "referred_user_id": c["_id"],
                                            "credited_at": {"$gte": datetime.combine(now.date(), datetime.min.time()).replace(tzinfo=timezone.utc)}
                                        })
                                        if not exists_ref:
                                            await db.referral_credits.insert_one({
                                                "user_id": referrer["_id"],
                                                "referred_user_id": c["_id"],
                                                "amount": amt_ref,
                                                "note": f"Referral commission from {c['name']} ({(ref_monthly_pct):.2f}% monthly)",
                                                "credited_at": now,
                                                "credited_by": "system",
                                            })
                                            total_ref += amt_ref
                                            ref_count += 1

                await db.system_state.update_one(
                    {"_id": "last_auto_credit"},
                    {"$set": {
                        "date": today,
                        "credited_clients": profit_count,
                        "total_profit": total_profit,
                        "credited_referrals": ref_count,
                        "total_referral": total_ref,
                        "at": now
                    }},
                    upsert=True,
                )
                logger.info("Auto-credit: %d clients ₹%.2f, %d referrals ₹%.2f", profit_count, total_profit, ref_count, total_ref)
        except Exception as e:
            logger.exception("Auto-credit error: %s", e)
        await asyncio.sleep(3600)  # check hourly


@app.on_event("shutdown")
async def shutdown():
    client.close()

